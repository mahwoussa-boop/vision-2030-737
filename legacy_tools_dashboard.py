# -*- coding: utf-8 -*-
"""
أدوات v11 المعزولة: المقارنة، مدقق المتجر، معالج SEO.
يستورد من legacy_core فقط (لا يمسّ engines/mahwous_core.py في v26).
"""
from __future__ import annotations

import io
import json
import logging
import os
import re
import sys
import time
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Union

import pandas as pd
import requests
import streamlit as st

try:
    from dotenv import load_dotenv

    load_dotenv()
except ImportError:
    pass

from legacy_core import (
    SALLA_SEO_COLS,
    StrictFilterOptions,
    legacy_apply_strict_pipeline_filters,
    format_salla_date_yyyy_mm_dd,
    normalize_price_digits,
    parse_price_numeric,
    validate_export_product_dataframe,
    validate_input_dataframe,
    validate_export_brands_list,
    validate_export_seo_dataframe,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from rapidfuzz import fuzz as rf_fuzz

    HAS_RAPIDFUZZ = True
except ImportError:
    HAS_RAPIDFUZZ = False

try:
    import anthropic

    HAS_ANTHROPIC = True
except ImportError:
    HAS_ANTHROPIC = False

# ══════════════════════════════════════════════════════════════════
#  مسارات البيانات — مجلد تطبيق v26
# ══════════════════════════════════════════════════════════════════
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")

MAHWOUS_SITE_BASE = os.environ.get("MAHWOUS_SITE_BASE", "https://mahwous.com").rstrip("/")
PUBLIC_APP_URL = os.environ.get(
    "PUBLIC_APP_URL",
    "https://mahwous-automation-production.up.railway.app",
).rstrip("/")

SALLA_COLS = [
    "النوع ", "أسم المنتج", "تصنيف المنتج", "صورة المنتج",
    "وصف صورة المنتج", "نوع المنتج", "سعر المنتج", "الوصف",
    "هل يتطلب شحن؟", "رمز المنتج sku", "سعر التكلفة", "السعر المخفض",
    "تاريخ بداية التخفيض", "تاريخ نهاية التخفيض",
    "اقصي كمية لكل عميل", "إخفاء خيار تحديد الكمية",
    "اضافة صورة عند الطلب", "الوزن", "وحدة الوزن",
    "الماركة", "العنوان الترويجي", "تثبيت المنتج",
    "الباركود", "السعرات الحرارية", "MPN", "GTIN",
    "خاضع للضريبة ؟", "سبب عدم الخضوع للضريبة",
    "[1] الاسم", "[1] النوع", "[1] القيمة", "[1] الصورة / اللون",
    "[2] الاسم", "[2] النوع", "[2] القيمة", "[2] الصورة / اللون",
    "[3] الاسم", "[3] النوع", "[3] القيمة", "[3] الصورة / اللون",
]

SALLA_BRANDS_COLS = [
    "اسم الماركة",
    "وصف مختصر عن الماركة",
    "صورة شعار الماركة",
    "(إختياري) صورة البانر",
    "(Page Title) عنوان صفحة العلامة التجارية",
    "(SEO Page URL) رابط صفحة العلامة التجارية",
    "(Page Description) وصف صفحة العلامة التجارية",
]

SALLA_PRICE_COLS = [
    "No.", "النوع ", "أسم المنتج", "رمز المنتج sku",
    "سعر المنتج", "سعر التكلفة", "السعر المخفض",
    "تاريخ بداية التخفيض", "تاريخ نهاية التخفيض",
]


def mahwous_brand_url(slug: str) -> str:
    s = str(slug or "").strip().strip("/")
    if not s:
        return f"{MAHWOUS_SITE_BASE}/brands"
    return f"{MAHWOUS_SITE_BASE}/brands/{s}"


def mahwous_category_url(path: str) -> str:
    p = str(path or "").strip().lstrip("/")
    return f"{MAHWOUS_SITE_BASE}/{p}" if p else MAHWOUS_SITE_BASE


def _get_secret(*keys: str) -> str:
    for k in keys:
        try:
            sec = getattr(st, "secrets", None)
            if sec is not None and k in sec:
                v = sec[k]
                if v is not None and str(v).strip():
                    return str(v).strip()
        except Exception:
            pass
        v = os.environ.get(k, "")
        if v and str(v).strip():
            return str(v).strip()
    return ""


def _effective_anthropic_api_key() -> str:
    try:
        k = str(st.session_state.get("api_key", "") or "").strip()
    except Exception:
        k = ""
    if k:
        return k
    return _get_secret("ANTHROPIC_API_KEY")


def configure_app_logging() -> logging.Logger:
    log = logging.getLogger("mahwous.legacy_ui")
    if log.handlers:
        return log
    log.setLevel(logging.INFO)
    try:
        os.makedirs(os.path.join(DATA_DIR, "logs"), exist_ok=True)
        fh = logging.FileHandler(
            os.path.join(DATA_DIR, "logs", "mahwous_legacy.log"),
            encoding="utf-8",
        )
        fh.setLevel(logging.INFO)
        fh.setFormatter(logging.Formatter(
            "%(asctime)s | %(levelname)s | %(name)s | %(message)s"
        ))
        log.addHandler(fh)
    except OSError:
        pass
    sh = logging.StreamHandler()
    sh.setLevel(logging.WARNING)
    sh.setFormatter(logging.Formatter("%(levelname)s | %(message)s"))
    log.addHandler(sh)
    return log


APP_LOG = configure_app_logging()


def anthropic_messages_create(client, **kwargs):
    last_err = None
    for attempt in range(3):
        try:
            return client.messages.create(**kwargs)
        except Exception as e:
            last_err = e
            err_l = str(e).lower()
            name = type(e).__name__
            retryable = (
                "429" in err_l
                or "rate limit" in err_l
                or "too many requests" in err_l
                or "rate_limit" in name.lower()
                or "overloaded" in err_l
            )
            if retryable and attempt < 2:
                time.sleep(2)
                APP_LOG.warning("anthropic retry %s/3 after %s: %s", attempt + 1, name, e)
                continue
            raise
    if last_err:
        raise last_err
    raise RuntimeError("anthropic_messages_create: unreachable")


def _parse_json_object_from_llm_text(raw: str, *, context: str = "") -> dict:
    empty: dict = {}
    if raw is None:
        APP_LOG.error("LLM JSON: raw is None (context=%s)", context)
        print(f"[mahwous] LLM JSON: empty raw (context={context!r})", file=sys.stderr)
        return empty
    s = str(raw).strip()
    if not s:
        APP_LOG.error("LLM JSON: empty string (context=%s)", context)
        print(f"[mahwous] LLM JSON: empty string (context={context!r})", file=sys.stderr)
        return empty
    if s.startswith("```"):
        lines = s.split("\n")
        if lines and lines[0].strip().startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        s = "\n".join(lines).strip()
    dec = json.JSONDecoder()
    positions = []
    pos = 0
    while len(positions) < 48:
        i = s.find("{", pos)
        if i < 0:
            break
        positions.append(i)
        pos = i + 1
    if not positions:
        APP_LOG.error("LLM JSON: no '{' in response (context=%s) snippet=%r", context, s[:400])
        print(f"[mahwous] LLM JSON: no '{{' (context={context!r})", file=sys.stderr)
        return empty
    last_err = None
    for start in positions:
        try:
            obj, _end = dec.raw_decode(s, start)
            if isinstance(obj, dict):
                return obj
            last_err = "root is not a JSON object"
        except json.JSONDecodeError as e:
            last_err = str(e)
            continue
    APP_LOG.error(
        "LLM JSON: unparseable after %d brace position(s) (context=%s) last_err=%s snippet=%r",
        len(positions),
        context,
        last_err,
        s[:500],
    )
    print(
        f"[mahwous] LLM JSON unparseable (context={context!r}): {last_err}",
        file=sys.stderr,
    )
    return empty


def _legacy_inject_css():
    """نفس ثيم التطبيق (config.toml: خلفية #0e1117، نص #fafafa) + لمسات v26 (Tajawal، بنفسجي)."""
    st.markdown(
        """
<style>
@import url('https://fonts.googleapis.com/css2?family=Tajawal:wght@400;700;900&display=swap');
/* توحيد مع بقية الصفحات — لا نغيّر خلفية .stApp؛ نعتمد theme.backgroundColor */
section.main .block-container,
section.main .block-container * {
  font-family: 'Tajawal', sans-serif !important;
}
section.main .block-container {
  max-width: 1400px !important;
  padding-top: 1rem !important;
}
section.main {
  color: var(--text-color, #fafafa) !important;
}
section.main [data-testid="stMarkdownContainer"] p,
section.main [data-testid="stMarkdownContainer"] li,
section.main .stMarkdown p {
  color: var(--text-color, #fafafa) !important;
}
section.main h1 {
  color: #e8e6ff !important;
  -webkit-text-fill-color: #e8e6ff !important;
  background: none !important;
  background-clip: unset !important;
  -webkit-background-clip: unset !important;
  font-weight: 900 !important;
  letter-spacing: -0.02em;
}
section.main h2, section.main h3 {
  color: var(--text-color, #fafafa) !important;
}
section.main .stCaption {
  color: rgba(250, 250, 250, 0.68) !important;
}
.al-info {
  background: #1A1A2E;
  border: 1px solid #333344;
  border-right: 4px solid #6C63FF;
  border-radius: 10px;
  padding: 12px 16px;
  font-size: 0.88rem;
  color: #c5c2f0;
  margin: 10px 0;
  direction: rtl;
  box-shadow: 0 4px 16px rgba(108, 99, 255, 0.08);
}
.al-ok {
  background: rgba(0, 200, 83, 0.12);
  border: 1px solid rgba(0, 200, 83, 0.35);
  border-right: 4px solid #00C853;
  border-radius: 10px;
  padding: 12px 16px;
  font-size: 0.88rem;
  color: #b9f6ca;
  margin: 10px 0;
  direction: rtl;
}
.sec-title {
  display: flex;
  align-items: center;
  gap: 10px;
  margin: 20px 0 12px;
  direction: rtl;
}
.sec-title .bar {
  width: 5px;
  height: 26px;
  border-radius: 4px;
  background: linear-gradient(180deg, #6C63FF, #5548d4);
  box-shadow: 0 0 12px rgba(108, 99, 255, 0.45);
}
.sec-title h3 {
  margin: 0;
  font-size: 1.08rem;
  font-weight: 800;
  color: #fafafa;
}
.stats-bar {
  display: flex;
  gap: 12px;
  flex-wrap: wrap;
  margin: 14px 0;
}
.stat-box {
  flex: 1;
  min-width: 110px;
  background: #1A1A2E;
  border-radius: 12px;
  padding: 14px 16px;
  text-align: center;
  border: 1px solid #333344;
  box-shadow: 0 4px 14px rgba(0, 0, 0, 0.12);
  transition: box-shadow 0.2s ease, border-color 0.2s ease;
}
.stat-box:hover {
  box-shadow: 0 4px 16px rgba(108, 99, 255, 0.15);
  border-color: #6C63FF66;
}
.stat-box .n {
  font-size: 1.9rem;
  font-weight: 900;
  color: #6C63FF;
  line-height: 1;
}
.stat-box .lb {
  font-size: 0.76rem;
  color: #8B8B8B;
  margin-top: 4px;
}
.upload-zone {
  border: 2px dashed #333344;
  border-radius: 14px;
  padding: 2.2rem;
  text-align: center;
  background: linear-gradient(135deg, #0e1117 0%, #1a1a2e 100%);
  color: #a8a4e0;
}
.upload-zone .uz-icon { font-size: 2.4rem; filter: drop-shadow(0 0 8px rgba(108,99,255,.4)); }
.upload-zone .uz-title { color: #e8e6ff !important; font-weight: 800; }
.upload-zone .uz-sub { color: #8B8B8B !important; }
.cmp-card {
  background: #1A1A2E;
  border: 1px solid #333344;
  border-radius: 12px;
  padding: 14px;
  margin-bottom: 12px;
  direction: rtl;
  color: #e0e0e8;
}
/* يبقى نص البطاقات الداكنة فاتحاً رغم قواعد المحتوى الفاتح أعلاه */
section.main .cmp-card,
section.main .cmp-card p,
section.main .cmp-card span {
  color: #e0e0e8 !important;
}
section.main .al-info,
section.main .al-info p,
section.main .al-info span {
  color: #c5c2f0 !important;
}
section.main .stat-box .n { color: #6C63FF !important; }
section.main .stat-box .lb { color: #8B8B8B !important; }
.cmp-card.suspect {
  border-color: #ff9800;
  background: linear-gradient(135deg, rgba(255, 152, 0, 0.08), #1A1A2E);
  box-shadow: 0 0 0 1px rgba(255, 152, 0, 0.2);
}
.gdiv {
  height: 1px;
  border: none;
  margin: 20px 0;
  background: linear-gradient(90deg, transparent, rgba(108, 99, 255, 0.35), transparent);
}
.prog-ok {
  background: rgba(0, 200, 83, 0.12);
  border-radius: 8px;
  padding: 8px 12px;
  margin: 4px 0;
  font-size: 0.82rem;
  color: #00C853;
  border: 1px solid rgba(0, 200, 83, 0.25);
}
.prog-run {
  background: rgba(255, 152, 0, 0.1);
  border-radius: 8px;
  padding: 8px 12px;
  margin: 4px 0;
  font-size: 0.82rem;
  color: #ffb74d;
  border: 1px solid rgba(255, 152, 0, 0.25);
}
/* تبويبات Streamlit — متوافقة مع الخلفية الداكنة */
.stTabs [data-baseweb="tab-list"] {
  background: rgba(26, 26, 46, 0.55);
  border-radius: 10px;
  padding: 6px;
  gap: 6px;
  border: 1px solid #333344;
}
.stTabs [data-baseweb="tab"] {
  font-family: 'Tajawal', sans-serif !important;
  font-weight: 700 !important;
  border-radius: 8px !important;
  color: rgba(250, 250, 250, 0.88) !important;
}
.stTabs [aria-selected="true"] {
  background: linear-gradient(135deg, #6C63FF, #5548d4) !important;
  color: #fff !important;
}
section.main div.stButton > button[kind="primary"] {
  background: linear-gradient(135deg, #6C63FF, #5548d4) !important;
  color: #fff !important;
  border: none !important;
  font-weight: 700 !important;
}
section.main div.stButton > button[kind="primary"]:hover {
  box-shadow: 0 4px 14px rgba(108, 99, 255, 0.45) !important;
  opacity: 0.95 !important;
}
section.main div.stButton > button[kind="secondary"] {
  border-color: #44445a !important;
  color: #fafafa !important;
  background-color: rgba(255, 255, 255, 0.06) !important;
}
section.main label,
section.main .stSelectbox label,
section.main .stSlider label {
  color: rgba(250, 250, 250, 0.9) !important;
}
</style>
""",
        unsafe_allow_html=True,
    )


# -*- coding: utf-8 -*-
"""تنفيذ أدوات v11 — يُدمج داخل legacy_tools_dashboard.py."""

def _legacy_init_state():
    defaults = {
        "api_key": _get_secret("ANTHROPIC_API_KEY"),
        "google_api": _get_secret("GOOGLE_API_KEY"),
        "google_cse": _get_secret("GOOGLE_CSE_ID"),
        "brands_df": None,
        "categories_df": None,
        "cmp_new_df": None,
        "cmp_store_df": None,
        "cmp_results": None,
        "cmp_approved": {},
        "cmp_edit_name": {},
        "cmp_from_pipe": False,
        "cmp_cfg": {},
        "seo_proc_tab_input_df": None,
        "seo_proc_tab_output_df": None,
        "audit_df": None,
        "audit_results": None,
        "audit_fixed_df": None,
        "pipe_session_brands": [],
        "new_brands": [],
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v
    if st.session_state.brands_df is None:
        p = os.path.join(DATA_DIR, "brands.csv")
        if os.path.exists(p):
            try:
                st.session_state.brands_df = pd.read_csv(p, encoding="utf-8-sig")
            except Exception:
                pass
    if st.session_state.categories_df is None:
        p = os.path.join(DATA_DIR, "categories.csv")
        if os.path.exists(p):
            try:
                st.session_state.categories_df = pd.read_csv(p, encoding="utf-8-sig")
            except Exception:
                pass


def _find_header_row_index(raw: pd.DataFrame, salla_2row: bool) -> int:
    markers_ar = (
        "أسم المنتج", "اسم المنتج", "نوع ", "no.", "no ", "رمز المنتج",
        "تصنيف المنتج", "صورة المنتج", "الماركة", "سعر المنتج", "الوصف",
    )
    markers_en = ("product name", "name", "sku", "price", "image", "title", "category")
    n = min(len(raw), 28)
    best_i = 1 if salla_2row else 0
    best_score = -1
    for i in range(n):
        row = raw.iloc[i]
        cells = [str(x).strip() for x in row.tolist() if pd.notna(x) and str(x).strip().lower() not in ("nan", "none", "")]
        if not cells:
            continue
        joined = " ".join(cells).lower().replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
        if len(cells) <= 2 and "بيانات المنتج" in joined and not any(m.lower() in joined for m in ("أسم", "اسم", "no", "sku")):
            continue
        sc = 0
        for m in markers_ar:
            ml = m.lower().replace("أ", "ا")
            if ml in joined or any(m in c or ml in c.lower() for c in cells):
                sc += 1
        for m in markers_en:
            if m in joined:
                sc += 1
        if sc > best_score:
            best_score = sc
            best_i = i
    if best_score >= 2:
        return best_i
    return 1 if salla_2row else 0


def read_file(f, salla_2row: bool = False) -> pd.DataFrame:
    name = f.name.lower()
    PREVIEW = 45
    hdr_fallback = 1 if salla_2row else 0
    try:
        if name.endswith((".xlsx", ".xlsm", ".xls")):
            f.seek(0)
            raw_preview = pd.read_excel(f, header=None, dtype=str, nrows=PREVIEW)
            hdr_idx = _find_header_row_index(raw_preview, salla_2row)
            f.seek(0)
            df = pd.read_excel(f, skiprows=list(range(hdr_idx)), header=0, dtype=str)
        else:
            df = None
            last_err = None
            for enc in ("utf-8-sig", "utf-8", "cp1256", "latin-1"):
                try:
                    f.seek(0)
                    raw_preview = pd.read_csv(f, header=None, encoding=enc, dtype=str, nrows=PREVIEW)
                    hdr_idx = _find_header_row_index(raw_preview, salla_2row)
                    f.seek(0)
                    df = pd.read_csv(f, skiprows=list(range(hdr_idx)), header=0, encoding=enc, dtype=str)
                    break
                except UnicodeDecodeError as e:
                    last_err = e
                    continue
            if df is None:
                APP_LOG.error("read_file CSV decode failed: %s", last_err)
                st.error("تعذّر فك ترميز ملف CSV.")
                return pd.DataFrame()
        df = df.dropna(how="all").reset_index(drop=True)
        df.columns = [str(c).strip() for c in df.columns]
        drop_u = [c for c in df.columns if str(c).lower().startswith("unnamed")
                  and df[c].fillna("").astype(str).str.strip().replace("nan", "").eq("").all()]
        if drop_u:
            df = df.drop(columns=drop_u, errors="ignore")
        return df
    except Exception as e:
        try:
            f.seek(0)
            if name.endswith((".xlsx", ".xlsm", ".xls")):
                df = pd.read_excel(f, header=hdr_fallback, dtype=str)
            else:
                for enc in ("utf-8-sig", "utf-8", "cp1256", "latin-1"):
                    try:
                        f.seek(0)
                        df = pd.read_csv(f, header=hdr_fallback, encoding=enc, dtype=str)
                        break
                    except UnicodeDecodeError:
                        continue
            df = df.dropna(how="all").reset_index(drop=True)
            df.columns = [str(c).strip() for c in df.columns]
            return df
        except Exception as e2:
            APP_LOG.exception("read_file fallback failed: %s | %s", e, e2)
            st.error(f"تعذّر قراءة الملف: {e2}")
            return pd.DataFrame()


def sanitize_salla_price_for_export(val) -> str:
    s = str(val or "").strip()
    if not s or s.lower() in ("nan", "none"):
        return ""
    s = re.sub(r"(?i)\b(?:sar|sr|usd|eur|aed)\b", " ", s)
    s = re.sub(r"(?:ريال|ر\.?\s*س)", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"[$€£¥]", "", s)
    return normalize_price_digits(s)


def compact_html_desc(html: str) -> str:
    if not html or not isinstance(html, str):
        return ""
    lines = [ln.strip() for ln in html.splitlines()]
    lines = [ln for ln in lines if ln]
    return "\n".join(lines)


def _is_unclear_column_name(col_name: str) -> bool:
    s = str(col_name).strip().lower()
    if not s:
        return True
    if s.startswith("unnamed"):
        return True
    if re.match(r"^column\.?\d+$", s):
        return True
    return False


def _infer_sample_role(series: pd.Series) -> str:
    vals = series.head(5).dropna().astype(str).str.strip()
    vals = [v for v in vals if v and v.lower() not in ("nan", "none", "")]
    if not vals:
        return "unknown"
    img_n = sum(1 for v in vals if v.startswith("http") or ("http" in v and "//" in v))
    pr_n = sum(1 for v in vals if re.search(r"\d", v) and len(v) < 48 and not v.startswith("http"))
    sku_n = sum(1 for v in vals if 2 <= len(v) <= 48 and re.match(r"^[\w\-\sA-Za-z٠-٩0-9]+$", v) and len(v.split()) <= 4)
    nm_n = sum(1 for v in vals if len(v) > 14 and (" " in v or "عطر" in v or "مل" in v or "bar" in v.lower()))
    scores = {"image": img_n, "price": pr_n, "sku": sku_n, "name": nm_n}
    best = max(scores, key=scores.get)
    if scores[best] >= max(2, len(vals) - 1):
        return best
    if scores[best] >= 1 and len(vals) == 1:
        return best
    return "unknown"


def _guess_role_from_keywords(keywords: list) -> str:
    for kw in keywords:
        kl = str(kw).lower()
        if any(x in kl for x in ("sku", "رمز", "barcode")):
            return "sku"
        if any(x in kl for x in ("صورة", "image", "src", "img", "w-full")):
            return "image"
        if any(x in kl for x in ("سعر", "price", "amount", "text-sm")):
            return "price"
        if any(x in kl for x in ("ماركة", "brand")):
            return "brand"
        if any(x in kl for x in ("وصف", "desc")):
            return "desc"
        if any(x in kl for x in ("اسم", "name", "منتج", "أسم", "product", "title")):
            return "name"
    return "name"


def auto_guess_col(cols, keywords: list, df: pd.DataFrame = None) -> str:
    col_list = [str(c) for c in cols]
    want = _guess_role_from_keywords(keywords)
    for kw in keywords:
        for c in col_list:
            if kw.lower() not in c.lower():
                continue
            if df is None or c not in df.columns:
                return c
            if _is_unclear_column_name(c):
                got = _infer_sample_role(df[c])
                if got not in ("unknown", want):
                    continue
            return c
    if df is not None and len(df) > 0:
        for c in col_list:
            if c not in df.columns:
                continue
            if _infer_sample_role(df[c]) == want:
                return c
    return "— لا يوجد —"


def _fuzzy_ratio(a: str, b: str) -> int:
    a, b = str(a).lower().strip(), str(b).lower().strip()
    if not a or not b:
        return 0
    if a == b:
        return 100
    if HAS_RAPIDFUZZ:
        return int(rf_fuzz.token_set_ratio(a, b))
    longer = max(len(a), len(b))
    matches = 0
    j = 0
    for ch in a:
        while j < len(b):
            if b[j] == ch:
                matches += 1
                j += 1
                break
            j += 1
    return int(matches / longer * 100) if longer else 0


_DISCOUNT_TAIL_RE = re.compile(
    r"(?:\s*[-–—]\s*\d+(?:[.,]\d+)?\s*%|\s+\d+(?:[.,]\d+)?\s*%)\s*$",
    re.IGNORECASE,
)


def strip_trailing_discount_label(text: str) -> str:
    if not text:
        return text
    s = str(text).strip()
    for _ in range(4):
        ns = _DISCOUNT_TAIL_RE.sub("", s).strip().strip("-–—").strip()
        if ns == s:
            break
        s = ns
    return s


def normalize_brand_name_v2(s: str) -> str:
    if not s:
        return ""
    t = unicodedata.normalize("NFKC", str(s).strip())
    for a in ("أ", "إ", "آ", "ٱ"):
        t = t.replace(a, "ا")
    t = t.replace("ة", "ه")
    t = re.sub(r"\s+", " ", t).strip()
    return t.lower()


def _strip_brand_name_edges(raw: object) -> str:
    return str(raw or "").strip().strip(' ,.-\'"')


def is_discount_like_brand(s: str) -> bool:
    if not s or not str(s).strip():
        return True
    t = str(s).strip().replace("٪", "%")
    if re.fullmatch(r"[\s\-–—]*\d+(?:[.,]\d+)?\s*%\s*", t, re.IGNORECASE):
        return True
    return False


def clean_brand_name(brand_raw: str) -> str:
    if not brand_raw:
        return ""
    b = _strip_brand_name_edges(brand_raw)
    if is_discount_like_brand(b):
        return ""
    if len(b.split()) > 3 or len(b) <= 2:
        return ""
    bad_words = [
        "تستر", "عطر", "شامبو", "بلسم", "لوشن", "مقوي", "مسكرة", "حقيبة",
        "بخاخ", "كريم", "زيت", "صابون", "جل", "معطر", "بودي", "مجموعة",
        "طقم", "عينة", "سمبل", "tester", "perfume", "منتج", "ميني", "mini",
    ]
    if any(w in b.lower() for w in bad_words):
        return ""
    return b


def match_brand(name: str) -> dict:
    if not str(name).strip():
        return {"name": "", "page_url": ""}
    nl = str(name).lower()
    for b in st.session_state.get("pipe_session_brands", []):
        raw = str(b.get("name", "") or "")
        if not raw:
            continue
        for part in re.split(r"\s*\|\s*", raw):
            p = part.strip().lower()
            if p and p in nl:
                return {"name": raw, "page_url": str(b.get("page_url", "") or "")}
    bdf = st.session_state.brands_df
    if bdf is None:
        return {"name": "", "page_url": ""}
    col0 = bdf.columns[0]
    for _, row in bdf.iterrows():
        raw = str(row[col0])
        for part in re.split(r"\s*\|\s*", raw):
            p = part.strip().lower()
            if p and p in nl:
                return {
                    "name": raw,
                    "page_url": str(row.get(
                        "(SEO Page URL) رابط صفحة العلامة التجارية", "") or ""),
                }
    return {"name": "", "page_url": ""}


def register_pipe_session_brand(display_name: str, page_url: str = "") -> None:
    if not str(display_name or "").strip():
        return
    low = str(display_name).strip().lower()
    if low in ("unknown", "غير محدد", "غير معروف"):
        return
    lst = st.session_state.setdefault("pipe_session_brands", [])
    norm = normalize_brand_name_v2(display_name)
    for x in lst:
        if normalize_brand_name_v2(x.get("name", "")) == norm:
            return
    lst.append({"name": str(display_name).strip(), "page_url": str(page_url or "").strip()})


def to_slug(text: str) -> str:
    out = ""
    for c in str(text).lower():
        if c.isascii() and c.isalnum():
            out += c
        elif c in " -_":
            out += "-"
    return re.sub(r"-+", "-", out).strip("-")


def _clean_brand_value_for_salla_output(raw: object) -> str:
    b = _strip_brand_name_edges(raw)
    if not b:
        return ""
    ar_part = b
    en_part = ""
    if "|" in b:
        ar_part, en_part = b.split("|", 1)
        ar_part = ar_part.strip()
        en_part = en_part.strip()
    ar_part = re.sub(r"^(تستر|عطر|طقم|مجموعة)\s+", " ", ar_part, flags=re.IGNORECASE)
    ar_part = re.sub(r"(?:\s|^)(تستر|عطر|طقم|مجموعة)(?:\s|$)", " ", ar_part, flags=re.IGNORECASE)
    ar_part = re.sub(r"(?:^|\s)مزيل\s+عرق(?:\s|$)", " ", ar_part, flags=re.IGNORECASE)
    ar_part = re.sub(r"\s+", " ", ar_part).strip()
    if en_part:
        en_part = re.sub(r"(?:^|\s)(tester|parfum|perfume)(?:\s|$)", " ", en_part, flags=re.IGNORECASE)
        en_part = re.sub(r"\s+", " ", en_part).strip()
    if en_part:
        if not ar_part:
            ar_part = _strip_brand_name_edges(raw).split("|", 1)[0].strip() if "|" in _strip_brand_name_edges(raw) else _strip_brand_name_edges(raw)
            ar_part = re.sub(r"^(تستر|عطر|طقم|مجموعة)\s+", "", ar_part, flags=re.IGNORECASE)
            ar_part = re.sub(r"\s+", " ", ar_part).strip()
        return f"{ar_part} | {en_part}".strip(" |")
    return ar_part


def generate_new_brand(brand_name: str) -> dict:
    brand_name = _strip_brand_name_edges(brand_name)
    if is_discount_like_brand(brand_name):
        return {
            "name": "",
            "page_url": "",
            "اسم الماركة": "",
            "وصف مختصر عن الماركة": "",
            "صورة شعار الماركة": "",
            "(إختياري) صورة البانر": "",
            "(Page Title) عنوان صفحة العلامة التجارية": "",
            "(SEO Page URL) رابط صفحة العلامة التجارية": "",
            "(Page Description) وصف صفحة العلامة التجارية": "",
        }
    key = st.session_state.api_key
    formatted_name = brand_name
    en_name = brand_name
    desc = f"علامة تجارية متخصصة في العطور الفاخرة - {brand_name}"
    if key and HAS_ANTHROPIC:
        try:
            client = anthropic.Anthropic(api_key=key)
            prompt = (
                f"أنت خبير علامات تجارية عالمية. ترجم ونسق الماركة '{brand_name}' بدقة عالية. "
                "التزم بصيغة JSON المغلقة بدون أي نصوص أو مقدمات خارجها: "
                '{"formatted_name": "الاسم بالعربي | الاسم بالانجليزي", '
                '"en_name": "English name only", '
                '"desc": "وصف جذاب 30 كلمة لمتجر مهووس"}'
            )
            msg = anthropic_messages_create(
                client,
                model="claude-3-haiku-20240307", max_tokens=250,
                messages=[{"role": "user", "content": prompt}]
            )
            raw = msg.content[0].text.strip()
            data = _parse_json_object_from_llm_text(raw, context="generate_new_brand")
            if data:
                formatted_name = data.get("formatted_name", brand_name)
                en_name = data.get("en_name", brand_name)
                desc = data.get("desc", desc)
        except Exception:
            pass
    slug = to_slug(en_name)
    display_name = formatted_name.split("|")[0].strip() if "|" in formatted_name else formatted_name
    out = {
        "name": display_name,
        "page_url": slug,
        "اسم الماركة": formatted_name,
        "وصف مختصر عن الماركة": desc,
        "صورة شعار الماركة": "",
        "(إختياري) صورة البانر": "",
        "(Page Title) عنوان صفحة العلامة التجارية": f"عطور {formatted_name} الأصلية | مهووس",
        "(SEO Page URL) رابط صفحة العلامة التجارية": slug,
        "(Page Description) وصف صفحة العلامة التجارية": f"تسوّق أحدث عطور {formatted_name} الأصلية الفاخرة بأسعار حصرية من متجر مهووس.",
    }
    register_pipe_session_brand(
        out.get("اسم الماركة", "") or "",
        out.get("(SEO Page URL) رابط صفحة العلامة التجارية", "") or "",
    )
    return out


def match_category(name: str, gender: str = "") -> str:
    t = (str(name) + " " + str(gender)).lower()
    if any(w in t for w in ["رجال", "للرجال", "men", "homme", "رجالي"]):
        return "العطور > عطور رجالية"
    if any(w in t for w in ["نساء", "للنساء", "women", "femme", "نسائي"]):
        return "العطور > عطور نسائية"
    return "العطور > عطور للجنسين"


# ── محرك مقارنة مبسّط (متوافق مع مخرجات واجهة v11) ─────────────────
def run_smart_comparison(
    new_df: pd.DataFrame,
    store_df: pd.DataFrame,
    new_name_col: str,
    store_name_col: str,
    new_sku_col: str = None,
    store_sku_col: str = None,
    new_img_col: str = None,
    t_dup: int = 88,
    t_near: int = 75,
    t_review: int = 55,
    brands_list: list = None,
    store_brand_col: str = None,
) -> pd.DataFrame:
    store_sku_set = set()
    for _, row in store_df.iterrows():
        sku = str(row.get(store_sku_col, "") or "").strip() if store_sku_col else ""
        if sku:
            store_sku_set.add(sku.lower())

    results = []
    for i, row in new_df.iterrows():
        new_name = str(row.get(new_name_col, "") or "").strip()
        new_sku = str(row.get(new_sku_col, "") or "").strip() if new_sku_col else ""
        new_img = str(row.get(new_img_col, "") or "").strip() if new_img_col else str(
            row.get("صورة المنتج", "") or ""
        ).strip()
        if not new_name or new_name.lower() in ("nan", "none", ""):
            continue
        new_name = strip_trailing_discount_label(new_name)
        if not str(new_name).strip():
            continue

        competitor_brand = ""
        if brands_list:
            nl = new_name.lower()
            for b in brands_list:
                if str(b).lower() in nl:
                    competitor_brand = b
                    break

        if new_sku and new_sku.lower() in store_sku_set:
            cb = clean_brand_name(competitor_brand or "")
            results.append({
                "الاسم الجديد": new_name, "SKU الجديد": new_sku,
                "الماركة": cb or "",
                "التصنيف": "العطور",
                "أقرب تطابق في المتجر": new_name, "نسبة التشابه": 100,
                "الحالة": "مكرر (SKU)", "سبب القرار": "تطابق SKU مباشر",
                "الإجراء": "حذف", "_idx": i, "_img": new_img,
            })
            continue

        best_score = 0.0
        best_store_name = ""
        for _, srow in store_df.iterrows():
            sn = str(srow.get(store_name_col, "") or "").strip()
            if not sn:
                continue
            sc = float(_fuzzy_ratio(new_name, sn))
            if HAS_RAPIDFUZZ:
                try:
                    sc = max(sc, float(rf_fuzz.token_sort_ratio(new_name, sn)))
                except Exception:
                    pass
            if sc > best_score:
                best_score = sc
                best_store_name = sn

        if best_score >= float(t_dup):
            status, action = "مكرر", "حذف"
            reason = f"تطابق ({best_score:.1f}%) — {best_store_name[:60]}"
        elif best_score >= float(t_near):
            status, action = "مشبوه", "مراجعة"
            reason = f"تشابه حرج ({best_score:.1f}%) — {best_store_name[:50] if best_store_name else '—'}"
        elif best_score > 0:
            status, action = "جديد", "اعتماد"
            reason = f"أقرب تشابه ({best_score:.1f}%) — غير كافٍ"
        else:
            status, action = "جديد", "اعتماد"
            reason = "جديد — لا يوجد في متجرنا"

        cb_disp = clean_brand_name(competitor_brand or "")
        results.append({
            "الاسم الجديد": new_name,
            "SKU الجديد": new_sku,
            "الماركة": cb_disp,
            "التصنيف": "العطور",
            "أقرب تطابق في المتجر": best_store_name,
            "نسبة التشابه": round(best_score, 1),
            "الحالة": status,
            "سبب القرار": reason,
            "الإجراء": action,
            "_idx": i,
            "_img": new_img,
        })

    return pd.DataFrame(results) if results else pd.DataFrame()


def dedupe_products_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty or "أسم المنتج" not in df.columns:
        return df
    sku_col = "رمز المنتج sku"
    name_col = "أسم المنتج"

    def _row_key(r) -> str:
        sku = str(r.get(sku_col, "") or "").strip().lower()
        if sku and sku not in ("nan", "none", ""):
            return "sku:" + sku
        t = str(r.get(name_col, "") or "").strip().lower()
        for a in ("أ", "إ", "آ"):
            t = t.replace(a, "ا")
        t = re.sub(r"\s+", " ", t)
        return t

    d2 = df.copy()
    d2["_dk"] = d2.apply(_row_key, axis=1)
    d2 = d2.drop_duplicates(subset=["_dk"], keep="first")
    d2 = d2.drop(columns=["_dk"], errors="ignore").reset_index(drop=True)
    if "No." in d2.columns:
        d2["No."] = [str(i + 1) for i in range(len(d2))]
    return d2


def _style_header_row(ws, row_num: int, cols: list, bg: str = "0F0E0D", fg: str = "B8933A"):
    for i, col in enumerate(cols, 1):
        c = ws.cell(row_num, i, col)
        c.font = Font(bold=True, color="FFFFFF" if bg != "E8D5B7" else "0F0E0D",
                       name="Cairo", size=9)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True, readingOrder=2)
        c.border = Border(bottom=Side(style="thin", color=fg))
    ws.row_dimensions[row_num].height = 30


def _prepare_salla_product_df_for_export(df: pd.DataFrame) -> pd.DataFrame:
    df = dedupe_products_df(df.copy())
    for col in SALLA_COLS:
        if col not in df.columns:
            df[col] = ""
    df = df.reindex(columns=list(SALLA_COLS))
    df["النوع "] = "منتج"
    df["نوع المنتج"] = "منتج جاهز"
    df["خاضع للضريبة ؟"] = "نعم"
    df["سعر المنتج"] = df["سعر المنتج"].apply(sanitize_salla_price_for_export)
    df["السعر المخفض"] = df["السعر المخفض"].apply(sanitize_salla_price_for_export)

    def _fix_sale_price_vs_regular(row: pd.Series) -> str:
        disc_s = str(row.get("السعر المخفض", "") or "").strip()
        if not disc_s or disc_s.lower() in ("nan", "none"):
            return ""
        ok_r, p_reg = parse_price_numeric(row.get("سعر المنتج", ""))
        ok_d, p_disc = parse_price_numeric(row.get("السعر المخفض", ""))
        if ok_r and ok_d and p_disc >= p_reg:
            return ""
        return disc_s

    df["السعر المخفض"] = df.apply(_fix_sale_price_vs_regular, axis=1)
    df["الماركة"] = df["الماركة"].apply(_clean_brand_value_for_salla_output)
    if "رمز المنتج sku" in df.columns:
        df["رمز المنتج sku"] = df.apply(
            lambda row: f"V-{row.name + 1}"
            if not str(row.get("رمز المنتج sku", "") or "").strip()
            else str(row["رمز المنتج sku"]),
            axis=1,
        )
    if "أسم المنتج" in df.columns:
        def _export_weight_from_name(n) -> str:
            s = str(n or "")
            if "طقم" in s or "مجموعة" in s:
                return "0.5"
            return "0.2"
        df["الوزن"] = df["أسم المنتج"].apply(_export_weight_from_name)
    df["وحدة الوزن"] = "kg"
    if "أسم المنتج" in df.columns:
        df["تصنيف المنتج"] = df["أسم المنتج"].apply(lambda n: match_category(str(n), ""))
    if "تصنيف المنتج" in df.columns:
        df["تصنيف المنتج"] = df["تصنيف المنتج"].astype(str).apply(
            lambda v: "العطور" if "العطور > عطور للجنسين" in v else v
        )

    def _export_cost_row(row) -> str:
        raw = row.get("سعر التكلفة", "")
        s = sanitize_salla_price_for_export(raw)
        if s:
            return s
        ok, price = parse_price_numeric(row.get("سعر المنتج", ""))
        if ok and price > 0:
            d = round(price * 0.70, 2)
            if abs(d - int(d)) < 1e-9:
                return str(int(d))
            ds = f"{d:.2f}".rstrip("0").rstrip(".")
            return ds
        return ""

    df["سعر التكلفة"] = df.apply(_export_cost_row, axis=1)

    def _disc_date(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        s = str(v).strip()
        if not s or s.lower() in ("nan", "none", "nat"):
            return ""
        return format_salla_date_yyyy_mm_dd(v) or ""

    df["تاريخ بداية التخفيض"] = df["تاريخ بداية التخفيض"].apply(_disc_date)
    df["تاريخ نهاية التخفيض"] = df["تاريخ نهاية التخفيض"].apply(_disc_date)
    if "وصف صورة المنتج" in df.columns:
        df["وصف صورة المنتج"] = (
            df["وصف صورة المنتج"]
            .astype(str)
            .str.replace(r"[^\w\s\u0600-\u06FF]", "", regex=True)
            .str.strip()
        )
    if "اقصي كمية لكل عميل" in df.columns:
        df["اقصي كمية لكل عميل"] = ""
    if "إخفاء خيار تحديد الكمية" in df.columns:
        df["إخفاء خيار تحديد الكمية"] = ""
    if "اضافة صورة عند الطلب" in df.columns:
        df["اضافة صورة عند الطلب"] = ""
    return df


def export_product_xlsx(df: pd.DataFrame) -> bytes:
    if df is None:
        df = pd.DataFrame(columns=SALLA_COLS)
    elif not df.empty:
        df = _prepare_salla_product_df_for_export(df)
    wb = Workbook()
    ws = wb.active
    ws.title = "Salla Products Template Sheet"
    ws.cell(1, 1, "بيانات المنتج")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(SALLA_COLS))
    c = ws.cell(1, 1)
    c.font = Font(bold=True, color="FFFFFF", name="Cairo", size=12)
    c.fill = PatternFill("solid", fgColor="0F0E0D")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    _style_header_row(ws, 2, SALLA_COLS, bg="E8D5B7", fg="B8933A")
    for ri, (_, row) in enumerate(df.iterrows(), 3):
        for ci, col in enumerate(SALLA_COLS, 1):
            v = str(row.get(col, "") if pd.notna(row.get(col, "")) else "")
            cell = ws.cell(ri, ci, v)
            cell.alignment = Alignment(horizontal="right", vertical="top",
                                      wrap_text=(col == "الوصف"), readingOrder=2)
            if ri % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="FAFAF8")
        ws.row_dimensions[ri].height = 18
    W = {
        "أسم المنتج": 45, "الوصف": 55, "تصنيف المنتج": 38,
        "صورة المنتج": 46, "الماركة": 24, "No.": 13,
        "وصف صورة المنتج": 36,
    }
    for i, col in enumerate(SALLA_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = W.get(col, 14)
    ws.freeze_panes = "A3"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_product_csv(df: pd.DataFrame) -> bytes:
    if df is None:
        df = pd.DataFrame(columns=SALLA_COLS)
    elif not df.empty:
        df = _prepare_salla_product_df_for_export(df)
    out = io.StringIO()
    out.write("بيانات المنتج" + "," * (len(SALLA_COLS) - 1) + "\n")
    out.write(",".join(SALLA_COLS) + "\n")
    for _, row in df.iterrows():
        vals = []
        for c in SALLA_COLS:
            v = str(row.get(c, "") if pd.notna(row.get(c, "")) else "")
            if c == "الوصف":
                v = '"' + v.replace('"', '""') + '"'
            elif any(x in v for x in [",", "\n", '"']):
                v = '"' + v.replace('"', '""') + '"'
            vals.append(v)
        out.write(",".join(vals) + "\n")
    return out.getvalue().encode("utf-8-sig")


def _enforce_salla_product_seo_limits(title: str, desc: str) -> tuple:
    t = str(title or "")
    d = str(desc or "")
    if len(t) > 60:
        t = t[:60]
    if len(d) > 160:
        d = d[:160]
    return t, d


def _append_sku_to_seo_slug(url: str, sku_suffix: str) -> str:
    u = str(url or "").strip()
    if not sku_suffix:
        return u
    suf_raw = str(sku_suffix).strip()
    suf = to_slug(suf_raw)
    if not suf:
        suf = re.sub(r"[^a-z0-9-]+", "-", suf_raw.lower()).strip("-")
    if not suf:
        return u
    if u.endswith(suf) or u.endswith("-" + suf):
        return u
    return f"{u}-{suf}".strip("-") if u else suf


def gen_seo(
    name: str,
    brand: dict,
    size: str,
    tester: bool,
    gender: str,
    sku_suffix: str = "",
    type_hint: str = "",
) -> dict:
    bname = brand.get("name", "")
    parts = re.split(r"\s*\|\s*", bname)
    ben = parts[-1].strip() if len(parts) > 1 else bname
    pref = "تستر" if tester else "عطر"
    title = f"{pref} {name} {size} | {ben}".strip()
    hint = f"{type_hint} {name} {bname}"
    if ("تستر" in hint or "tester" in hint.lower()) and "تستر" not in title:
        title = f"تستر {title}".strip()
    desc = (f"تسوق {pref} {name} {size} الأصلي من {bname}. "
            f"عطر {gender} فاخر ثابت. أصلي 100% من مهووس.")
    slug = to_slug(f"{ben}-{name}-{size}".replace("مل", "ml"))
    slug = _append_sku_to_seo_slug(slug, sku_suffix)
    title, desc = _enforce_salla_product_seo_limits(title, desc)
    return {
        "url": slug,
        "title": title,
        "desc": desc,
        "alt": f"زجاجة {pref} {name} {size} الأصلية",
    }


def fetch_image(name: str, tester: bool = False) -> str:
    gk = st.session_state.google_api
    cx = st.session_state.google_cse
    if not gk or not cx:
        return ""
    try:
        q = name + (" tester box" if tester else " perfume bottle")
        r = requests.get(
            "https://www.googleapis.com/customsearch/v1",
            params={"key": gk, "cx": cx, "q": q,
                    "searchType": "image", "num": 1, "imgSize": "large"},
            timeout=10,
        )
        items = r.json().get("items", [])
        return items[0]["link"] if items else ""
    except Exception:
        return ""


def _ai_fetch_notes_only(name: str, brand_name: str, api_key: str) -> dict:
    try:
        client = anthropic.Anthropic(api_key=api_key)
        prompt = (
            f"أنت خبير كيميائي للعطور. استخرج المكونات الحقيقية لعطر '{name}' "
            f"من ماركة '{brand_name}'. "
            "لا تهلوس مكونات غير موجودة — إذا لم تعرفها اكتب 'غير متوفر'. "
            "الرد يجب أن يكون JSON مغلق بدون أي مقدمات أو نصوص خارجه:\n"
            '{"top": "مكونات القمة", "heart": "مكونات القلب", '
            '"base": "مكونات القاعدة", "family": "العائلة العطرية", "year": "سنة الإصدار"}'
        )
        msg = anthropic_messages_create(
            client,
            model="claude-3-haiku-20240307",
            max_tokens=300,
            temperature=0.1,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = msg.content[0].text.strip()
        parsed = _parse_json_object_from_llm_text(raw, context="ai_fetch_notes_only")
        if parsed:
            return parsed
    except Exception:
        pass
    return {"top": "غير متوفر", "heart": "غير متوفر", "base": "غير متوفر",
            "family": "غير متوفر", "year": "غير معروف"}


def _build_html_description(name: str, tester: bool, brand: dict,
                            size: str, gender: str, conc: str,
                            notes: dict) -> str:
    ptype = "تستر" if tester else "عطر"
    brand_name = brand.get("name", "غير محدد")
    brand_url = brand.get("page_url", "")
    brand_link = (
        f"<a href='{mahwous_brand_url(brand_url)}' target='_blank' rel='noopener'>{brand_name}</a>"
        if brand_url else brand_name
    )
    top = notes.get("top", "برغموت، ليمون")
    heart = notes.get("heart", "ورد، ياسمين")
    base = notes.get("base", "مسك، عنبر")
    family = notes.get("family", "عطرية")
    year = notes.get("year", "")
    gender_txt = ("للنساء" if "نساء" in gender
                  else "للرجال" if "رجال" in gender else "للجنسين")
    h = []
    h.append(f'<h2>{ptype} {brand_name} {name} {conc} {size} {gender_txt}</h2>')
    h.append(f'<p>اكتشف سحر <strong>{name}</strong> من <strong>{brand_link}</strong> — '
             f'عطر فاخر يجمع بين الأصالة والتميز. '
             f'صمّم خصيصاً {gender_txt} ليرسم بصمتك العطري بثقة وأناقة. '
             f'متوفّر بحجم {size} بتركيز <strong>{conc}</strong> لضمان ثبات استثنائي.</p>')
    h.append('<h3>تفاصيل المنتج</h3><ul>')
    h.append(f'<li><strong>الماركة:</strong> {brand_link}</li>')
    h.append(f'<li><strong>الاسم:</strong> {name}</li>')
    h.append(f'<li><strong>الجنس:</strong> {gender_txt}</li>')
    if family and "غير متوفر" not in str(family):
        h.append(f'<li><strong>العائلة العطرية:</strong> {family}</li>')
    h.append(f'<li><strong>الحجم:</strong> {size}</li>')
    h.append(f'<li><strong>التركيز:</strong> {conc}</li>')
    if year and year != "غير معروف":
        h.append(f'<li><strong>سنة الإصدار:</strong> {year}</li>')
    h.append(f'<li><strong>نوع المنتج:</strong> {"تستر (Tester)" if tester else "عطر أصلي"}</li>')
    h.append('</ul>')
    h.append('<h3>رحلة العطر — الهرم العطري</h3>')
    h.append(f'<p>يأخذك <strong>{name}</strong> في رحلة عطرية متكاملة.</p><ul>')
    h.append(f'<li><strong>المقدمة:</strong> {top}</li>')
    h.append(f'<li><strong>القلب:</strong> {heart}</li>')
    h.append(f'<li><strong>القاعدة:</strong> {base}</li></ul>')
    h.append(f'<p><strong>عالمك العطري يبدأ من مهووس.</strong> أصلي 100% | شحن سريع داخل السعودية.</p>')
    return "\n".join(h)


def ai_generate(name: str, tester: bool, brand: dict,
                size: str, gender: str, conc: str) -> str:
    key = st.session_state.api_key
    if not key:
        return "<p>أضف مفتاح Anthropic API في الإعدادات أو secrets (ANTHROPIC_API_KEY).</p>"
    if not HAS_ANTHROPIC:
        return "<p>تعذّر تحميل مكتبة anthropic.</p>"
    notes = _ai_fetch_notes_only(name, brand.get("name", ""), key)
    return _build_html_description(name, tester, brand, size, gender, conc, notes)


def build_empty_salla_row() -> dict:
    r = {c: "" for c in SALLA_COLS}
    r["النوع "] = "منتج"
    r["نوع المنتج"] = "منتج جاهز"
    r["هل يتطلب شحن؟"] = "نعم"
    r["خاضع للضريبة ؟"] = "نعم"
    r["الوزن"] = "0.2"
    r["وحدة الوزن"] = "kg"
    r["حالة المنتج"] = "مرئي"
    r["اقصي كمية لكل عميل"] = "0"
    r["إخفاء خيار تحديد الكمية"] = "0"
    r["اضافة صورة عند الطلب"] = "0"
    return r


def fill_row(name, price="", sku="", image="", desc="",
             brand=None, category="", seo=None, no="",
             weight="0.2", weight_unit="kg", size=""):
    if brand is None:
        brand = {}
    if seo is None:
        seo = {}
    r = build_empty_salla_row()
    r["No."] = str(no)
    r["أسم المنتج"] = str(name)
    r["سعر المنتج"] = normalize_price_digits(price)
    r["رمز المنتج sku"] = str(sku)
    r["صورة المنتج"] = str(image)
    r["وصف صورة المنتج"] = seo.get("alt", "")
    r["الوصف"] = compact_html_desc(str(desc))
    r["الماركة"] = _clean_brand_value_for_salla_output(brand.get("name", ""))
    r["تصنيف المنتج"] = str(category)
    r["الوزن"] = str(weight) if weight else "0.2"
    r["وحدة الوزن"] = str(weight_unit) if weight_unit else "kg"
    if not str(price).strip() or str(price).strip() in ("0", "nan", "None"):
        r["اقصي كمية لكل عميل"] = "0"
    return r


def _normalize_product_size_ml(size: str) -> str:
    s = str(size or "").strip()
    if not s or s.lower() in ("nan", "none"):
        return ""
    low = s.lower()
    if "مل" in s or re.search(r"\bml\b", low):
        num_m = re.search(r"(\d+(?:[.,]\d+)?)", s.replace("٫", ".").replace(",", "."))
        if num_m:
            val = float(num_m.group(1).replace(",", "."))
            if abs(val - round(val)) < 1e-9:
                return f"{int(round(val))} مل"
            t = f"{val:.4f}".rstrip("0").rstrip(".")
            return f"{t} مل"
        return s
    t = s.replace("٫", ".").replace(",", ".").strip()
    m = re.fullmatch(r"(\d+\.?\d*|\.\d+)", t)
    if not m:
        return s
    val = float(m.group(1))
    if 0 < val < 1:
        val = val * 100.0
    if abs(val - round(val)) < 1e-9:
        return f"{int(round(val))} مل"
    t2 = f"{val:.4f}".rstrip("0").rstrip(".")
    return f"{t2} مل"


def extract_product_attrs(name: str) -> dict:
    return {
        "size": 100.0,
        "type": "عطر تجاري",
        "concentration": "EDP",
        "clean_name": str(name or ""),
        "core_name": str(name or ""),
        "brand": "",
        "category": "العطور",
    }


def standardize_product_name(raw_name: str, brand_name: str) -> str:
    b = str(brand_name or "").split("|")[0].strip()
    if b and b not in raw_name:
        return f"{str(raw_name).strip()} {b}".strip()
    return str(raw_name).strip()


def generate_seo_data_ai(product_name: str, missing_fields: list[str]) -> dict:
    empty_out = {}
    api_key = getattr(st.session_state, "api_key", None)
    if not api_key or not HAS_ANTHROPIC:
        return empty_out
    try:
        product_name = str(product_name or "").strip()
        if not product_name:
            return empty_out
        miss = {x for x in (missing_fields or []) if x in ("url", "title", "desc")}
        if not miss:
            return empty_out
        rules = []
        if "url" in miss:
            rules.append("SEO Page URL: English lowercase, hyphen-separated.")
        if "title" in miss:
            rules.append("SEO Page Title: Arabic, max 60 chars, include product + مهووس.")
        if "desc" in miss:
            rules.append("SEO Page Description: Arabic, max 160 chars, CTA.")
        prompt = (
            "أنت خبير SEO لمتجر عطور مهووس.\n"
            f"اسم المنتج: {product_name}\n"
            f"missing_fields: {sorted(list(miss))}\n\n"
            + "\n".join(f"- {r}" for r in rules)
            + "\n\nأعد JSON فقط: url_slug, page_title, meta_description (حسب المطلوب فقط).\n"
        )
        client = anthropic.Anthropic(api_key=api_key)
        msg = anthropic_messages_create(
            client,
            model="claude-3-haiku-20240307",
            max_tokens=500,
            temperature=0.1,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = msg.content[0].text.strip()
        data = _parse_json_object_from_llm_text(raw, context="generate_seo_data_ai")
        if not data:
            return empty_out
        out = {}
        if "url_slug" in data:
            out["url_slug"] = str(data.get("url_slug", "")).strip()
        if "page_title" in data:
            out["page_title"] = str(data.get("page_title", "")).strip()
        if "meta_description" in data:
            out["meta_description"] = str(data.get("meta_description", "")).strip()
        if out:
            t_lim, d_lim = _enforce_salla_product_seo_limits(
                out.get("page_title", ""), out.get("meta_description", "")
            )
            if "page_title" in out:
                out["page_title"] = t_lim
            if "meta_description" in out:
                out["meta_description"] = d_lim
        return out
    except Exception:
        return empty_out


def _pick_first_non_http_text_col(df: pd.DataFrame) -> str:
    txt_cols = [c for c in df.columns if pd.api.types.is_object_dtype(df[c]) or pd.api.types.is_string_dtype(df[c])]
    for c in txt_cols:
        s = df[c].fillna("").astype(str).str.strip()
        nz = s[s != ""]
        if nz.empty:
            continue
        http_ratio = nz.str.contains(r"https?://", case=False, regex=True).mean()
        if http_ratio < 0.5:
            return c
    return txt_cols[0] if txt_cols else ""


def _guess_competitor_name_col(df: pd.DataFrame) -> str:
    if "أسم المنتج" in df.columns:
        return "أسم المنتج"
    g = auto_guess_col(
        df.columns,
        ["أسم المنتج", "اسم المنتج", "name", "title", "product", "styles_productcard", "styles_productCard"],
        df,
    )
    if g and g != "— لا يوجد —":
        return g
    return _pick_first_non_http_text_col(df)


def _guess_competitor_price_col(df: pd.DataFrame) -> str:
    if "سعر المنتج" in df.columns:
        return "سعر المنتج"
    g = auto_guess_col(
        df.columns,
        ["سعر المنتج", "السعر", "price", "text-sm", "amount", "value"],
        df,
    )
    if g and g != "— لا يوجد —":
        return g
    return ""


def _read_competitor_file(df_raw: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    if df_raw is None or df_raw.empty:
        return pd.DataFrame(), {"name_col": "", "price_col": "", "rows_before": 0, "rows_after": 0}
    df = df_raw.copy()
    name_col = _guess_competitor_name_col(df)
    price_col = _guess_competitor_price_col(df)
    if name_col and name_col != "أسم المنتج" and "أسم المنتج" not in df.columns:
        df = df.rename(columns={name_col: "أسم المنتج"})
    if price_col and price_col != "سعر المنتج" and "سعر المنتج" not in df.columns:
        df = df.rename(columns={price_col: "سعر المنتج"})
    if "أسم المنتج" not in df.columns:
        fb = _pick_first_non_http_text_col(df)
        if fb:
            df = df.rename(columns={fb: "أسم المنتج"})
    df = df.dropna(how="all").reset_index(drop=True)
    return df, {
        "name_col": "أسم المنتج" if "أسم المنتج" in df.columns else name_col,
        "price_col": "سعر المنتج" if "سعر المنتج" in df.columns else price_col,
        "rows_before": len(df_raw),
        "rows_after": len(df),
    }


def apply_competitor_exclusions(
    df: pd.DataFrame,
    name_col: str = "أسم المنتج",
    *,
    exclude_samples: bool = True,
    exclude_makeup: bool = True,
    exclude_accessories: bool = True,
    exclude_missing_sizes: bool = True,
) -> tuple[pd.DataFrame, dict]:
    if df is None or df.empty or name_col not in df.columns:
        return df, {"input_rows": 0, "output_rows": 0}
    work = df.copy()
    names = work[name_col].fillna("").astype(str)
    nl = names.str.lower()
    is_tester = nl.str.contains(r"(?:\btester\b|تستر|تيستر)", regex=True, na=False)
    sample_kw = r"(?:عينة|سمبل|فايال|\bvial\b|\bsample\b)"
    sample_size = r"\b[1-8]\s*(?:مل|ml)\b"
    makeup_kw = r"(?:مكياج|ارواج|روج|ماسكارا|فاونديشن|ظل|بلشر|كونسيلر)"
    acc_kw = r"(?:شنطة|كيس|تغليف|كرتون|مبخرة|فحم|ميدالية)"
    has_size = nl.str.contains(r"(?:مل|ml|oz|غرام|جرام)", regex=True, na=False)
    drop_sample = (nl.str.contains(sample_kw, regex=True, na=False) | nl.str.contains(sample_size, regex=True, na=False)) & (~is_tester)
    drop_makeup = nl.str.contains(makeup_kw, regex=True, na=False)
    drop_acc = nl.str.contains(acc_kw, regex=True, na=False)
    drop_no_size = ~has_size
    mask = pd.Series(False, index=work.index)
    stats = {"input_rows": len(work), "dropped_samples": 0, "dropped_makeup": 0, "dropped_accessories": 0, "dropped_missing_sizes": 0}
    if exclude_samples:
        mask |= drop_sample
        stats["dropped_samples"] = int(drop_sample.sum())
    if exclude_makeup:
        mask |= drop_makeup
        stats["dropped_makeup"] = int(drop_makeup.sum())
    if exclude_accessories:
        mask |= drop_acc
        stats["dropped_accessories"] = int(drop_acc.sum())
    if exclude_missing_sizes:
        mask |= drop_no_size
        stats["dropped_missing_sizes"] = int(drop_no_size.sum())
    out = work.loc[~mask].copy()
    stats["output_rows"] = len(out)
    return out, stats


def render_compare_tab():
    _legacy_init_state()
    _legacy_inject_css()
    st.markdown("""<div class="al-info">
    قارن ملف المنتجات الجديدة بملف المتجر الأساسي. تُعرض المنتجات <strong>المشبوهة</strong> في كروت
    بصورتين للمراجعة السريعة.
    </div>""", unsafe_allow_html=True)

    if st.session_state.get("cmp_from_pipe") and st.session_state.cmp_new_df is not None:
        st.markdown('<div class="al-ok">📎 تم تحميل بيانات من جلسة سابقة.</div>', unsafe_allow_html=True)

    c_up1, c_up2 = st.columns(2)
    with c_up1:
        st.markdown("**ملف المنتجات الجديدة**")
        up_n = st.file_uploader("CSV / Excel", type=["csv", "xlsx", "xls"], key="cmp_up_new", label_visibility="collapsed")
        if up_n:
            dfn = read_file(up_n, salla_2row=True)
            if dfn.empty:
                dfn = read_file(up_n, salla_2row=False)
            if not dfn.empty:
                dfn_norm, map_info = _read_competitor_file(dfn)
                st.session_state.cmp_new_df = dfn_norm
                st.session_state.cmp_from_pipe = False
                st.success(f"✅ {len(dfn_norm)} منتج")
                if map_info.get("name_col") != "أسم المنتج":
                    st.caption(f"عمود الاسم: `{map_info.get('name_col', '')}`")
    with c_up2:
        st.markdown("**ملف المتجر الأساسي**")
        up_s = st.file_uploader("CSV / Excel", type=["csv", "xlsx", "xls"], key="cmp_up_store", label_visibility="collapsed")
        if up_s:
            dfs = read_file(up_s, salla_2row=True)
            if dfs.empty:
                dfs = read_file(up_s, salla_2row=False)
            if not dfs.empty:
                st.session_state.cmp_store_df = dfs
                st.success(f"✅ {len(dfs)} منتج في المتجر")

    if st.session_state.cmp_new_df is not None and st.session_state.cmp_store_df is not None:
        new_df = st.session_state.cmp_new_df
        store_df = st.session_state.cmp_store_df
        NONE_C = "— لا يوجد —"
        n_opts = [NONE_C] + list(new_df.columns)
        s_opts = [NONE_C] + list(store_df.columns)

        def _gi(cols, kws, df_, opts):
            g = auto_guess_col(cols, kws, df_)
            return opts.index(g) if g in opts else 0

        r1, r2, r3, r4 = st.columns(4)
        with r1:
            new_nm = st.selectbox("عمود الاسم (جديد):", n_opts,
                                  index=_gi(new_df.columns, ["اسم", "name", "منتج"], new_df, n_opts), key="cmp_nn")
        with r2:
            new_sk = st.selectbox("عمود SKU (جديد):", n_opts,
                                  index=_gi(new_df.columns, ["sku", "رمز"], new_df, n_opts), key="cmp_nsk")
        with r3:
            st_nm = st.selectbox("عمود الاسم (متجر):", s_opts,
                                 index=_gi(store_df.columns, ["اسم", "name", "منتج"], store_df, s_opts), key="cmp_sn")
        with r4:
            st_sk = st.selectbox("عمود SKU (متجر):", s_opts,
                                 index=_gi(store_df.columns, ["sku", "رمز"], store_df, s_opts), key="cmp_ssk")

        new_img_g = auto_guess_col(new_df.columns, ["صورة", "image", "src"], new_df)
        new_img_g = None if new_img_g == "— لا يوجد —" else new_img_g
        sim_thr = st.slider("عتبة التشابه للمشبوه (%):", 50, 95, 75, key="cmp_sim")
        fx1, fx2, fx3, fx4 = st.columns(4)
        with fx1:
            cmp_fx_samples = st.checkbox("استبعاد العينات", value=True, key="cmp_fx_samples")
        with fx2:
            cmp_fx_makeup = st.checkbox("استبعاد المكياج", value=True, key="cmp_fx_makeup")
        with fx3:
            cmp_fx_accessories = st.checkbox("استبعاد الكماليات", value=True, key="cmp_fx_accessories")
        with fx4:
            cmp_fx_missing_size = st.checkbox("استبعاد منتجات بدون حجم", value=True, key="cmp_fx_missing_size")

        if st.button("🔍 تشغيل المقارنة والعرض المرئي", type="primary", key="cmp_run", width="stretch"):
            if new_nm == NONE_C or st_nm == NONE_C:
                st.error("حدد عمود الاسم للملفين.")
            else:
                comp_for_compare = new_df.copy()
                if "أسم المنتج" not in comp_for_compare.columns and new_nm in comp_for_compare.columns:
                    comp_for_compare = comp_for_compare.rename(columns={new_nm: "أسم المنتج"})
                    new_nm = "أسم المنتج"
                comp_for_compare, ex_stats = apply_competitor_exclusions(
                    comp_for_compare,
                    name_col="أسم المنتج" if "أسم المنتج" in comp_for_compare.columns else new_nm,
                    exclude_samples=cmp_fx_samples,
                    exclude_makeup=cmp_fx_makeup,
                    exclude_accessories=cmp_fx_accessories,
                    exclude_missing_sizes=cmp_fx_missing_size,
                )
                if comp_for_compare.empty:
                    st.error("لا توجد منتجات صالحة بعد فلاتر الاستبعاد.")
                    st.stop()
                brands_l = []
                if st.session_state.brands_df is not None:
                    brands_l = (st.session_state.brands_df[st.session_state.brands_df.columns[0]]
                                .dropna().astype(str).str.strip().tolist())
                with st.spinner("جاري المقارنة..."):
                    res_df = run_smart_comparison(
                        new_df=comp_for_compare,
                        store_df=store_df,
                        new_name_col=new_nm,
                        store_name_col=st_nm,
                        new_sku_col=new_sk if new_sk != NONE_C else None,
                        store_sku_col=st_sk if st_sk != NONE_C else None,
                        new_img_col=new_img_g,
                        t_dup=88, t_near=sim_thr, t_review=50,
                        brands_list=brands_l,
                    )
                    st.session_state.cmp_results = res_df
                    st.session_state.cmp_cfg = {
                        "new_nm": new_nm, "st_nm": st_nm, "new_img": new_img_g,
                        "cmp_ex_stats": ex_stats,
                    }
                    sus = res_df[res_df["الحالة"] == "مشبوه"]
                    st.session_state.cmp_approved = {
                        int(r["_idx"]): True for _, r in sus.iterrows()
                    }
                st.rerun()

    if st.session_state.cmp_results is not None:
        res = st.session_state.cmp_results
        new_df = st.session_state.cmp_new_df
        store_df = st.session_state.cmp_store_df
        cfg = st.session_state.get("cmp_cfg", {})
        st_nm = cfg.get("st_nm", "أسم المنتج" if "أسم المنتج" in store_df.columns else store_df.columns[0])
        new_img_col = cfg.get("new_img")
        s_img_col = auto_guess_col(store_df.columns, ["صورة", "image", "src"], store_df)
        if s_img_col == "— لا يوجد —":
            s_img_col = None

        exact_dup = res[res["الحالة"].astype(str).str.contains("مكرر", na=False)]
        suspect = res[res["الحالة"] == "مشبوه"]
        new_clean = res[res["الحالة"] == "جديد"]

        st.markdown(f"""
        <div class="stats-bar">
          <div class="stat-box"><div class="n">{len(res)}</div><div class="lb">إجمالي</div></div>
          <div class="stat-box"><div class="n" style="color:#e53935">{len(exact_dup)}</div><div class="lb">مكرر</div></div>
          <div class="stat-box"><div class="n" style="color:#f9a825">{len(suspect)}</div><div class="lb">مشبوه</div></div>
          <div class="stat-box"><div class="n" style="color:#43a047">{len(new_clean)}</div><div class="lb">جديد</div></div>
        </div>
        """, unsafe_allow_html=True)
        exs = cfg.get("cmp_ex_stats", {})
        if exs:
            st.caption(
                f"فلترة المنافسين: دخل {exs.get('input_rows', 0)} → خرج {exs.get('output_rows', 0)} | "
                f"عينات: {exs.get('dropped_samples', 0)}، مكياج: {exs.get('dropped_makeup', 0)}، "
                f"كماليات: {exs.get('dropped_accessories', 0)}، بدون حجم: {exs.get('dropped_missing_sizes', 0)}"
            )

        if not suspect.empty:
            st.markdown("""<div class="sec-title"><div class="bar"></div>
            <h3>منتجات مشبوهة — مقارنة بصرية</h3></div>""", unsafe_allow_html=True)
            for _, srow in suspect.iterrows():
                idx = int(srow["_idx"])
                new_img_u = str(srow.get("_img", "") or "").split(",")[0].strip().replace(" ", "%20")
                if not new_img_u.startswith("http"):
                    new_img_u = ""
                store_match = str(srow.get("أقرب تطابق في المتجر", "") or "")
                store_img_u = ""
                if store_match and st_nm in store_df.columns:
                    try:
                        sm = store_df[store_df[st_nm].astype(str) == store_match]
                        if not sm.empty and s_img_col and s_img_col in store_df.columns:
                            store_img_u = str(sm.iloc[0].get(s_img_col, "") or "").split(",")[0].strip().replace(" ", "%20")
                            if not store_img_u.startswith("http"):
                                store_img_u = ""
                    except Exception:
                        pass

                ph = "width:120px;height:120px;background:#eee;border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:2rem"
                img_new = f'<img src="{new_img_u}" style="width:120px;height:120px;object-fit:cover;border-radius:10px" onerror="this.style.display=\'none\'">' if new_img_u else f'<div style="{ph}">🖼</div>'
                img_st = f'<img src="{store_img_u}" style="width:120px;height:120px;object-fit:cover;border-radius:10px" onerror="this.style.display=\'none\'">' if store_img_u else f'<div style="{ph}">🏪</div>'

                st.markdown('<div class="cmp-card suspect">', unsafe_allow_html=True)
                h1, h2, h3 = st.columns([2, 2, 1])
                with h1:
                    st.markdown("**🆕 جديد**", unsafe_allow_html=True)
                    st.markdown(img_new, unsafe_allow_html=True)
                    st.caption(srow.get("الاسم الجديد", ""))
                with h2:
                    st.markdown("**🏪 المتجر**", unsafe_allow_html=True)
                    st.markdown(img_st, unsafe_allow_html=True)
                    st.caption(store_match or "—")
                with h3:
                    st.metric("تشابه", f"{srow.get('نسبة التشابه', 0)}%")

                _ek = f"cmp_edit_{idx}"
                if _ek not in st.session_state:
                    st.session_state[_ek] = str(
                        st.session_state.cmp_edit_name.get(idx, srow.get("الاسم الجديد", ""))
                    )
                st.text_input("تعديل الاسم إن لزم", key=_ek)
                st.session_state.cmp_edit_name[idx] = st.session_state[_ek]

                b1, b2, b3 = st.columns(3)
                with b1:
                    if st.button("✅ اعتماد كجديد", key=f"cmp_ok_{idx}", width="stretch"):
                        st.session_state.cmp_approved[idx] = True
                        st.rerun()
                with b2:
                    if st.button("⛔ تجاهل (مكرر)", key=f"cmp_no_{idx}", width="stretch"):
                        st.session_state.cmp_approved[idx] = False
                        st.rerun()
                with b3:
                    if st.button("💾 حفظ التعديل على الاسم", key=f"cmp_sv_{idx}", width="stretch"):
                        st.toast("تم حفظ الاسم في المعاينة — اضغط اعتماد أو تجاهل")
                st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
        <h3>تصدير القائمة بعد القرار</h3></div>""", unsafe_allow_html=True)
        if st.button("⚡ بناء ملف المنتجات النهائي", type="primary", key="cmp_build"):
            new_src = st.session_state.cmp_new_df
            rows_out = []
            for _, rrow in res.iterrows():
                idx = int(rrow["_idx"])
                stt = str(rrow["الحالة"])
                if stt == "جديد":
                    if idx in new_src.index:
                        row_ser = new_src.loc[idx].copy()
                        if "الماركة" in rrow.index:
                            row_ser["الماركة"] = rrow.get("الماركة", "")
                        rows_out.append(row_ser)
                elif stt == "مشبوه":
                    ap = st.session_state.cmp_approved.get(idx, True)
                    if ap and idx in new_src.index:
                        row_ser = new_src.loc[idx].copy()
                        if idx in st.session_state.cmp_edit_name:
                            if "أسم المنتج" in row_ser.index:
                                row_ser["أسم المنتج"] = st.session_state.cmp_edit_name[idx]
                        if "الماركة" in rrow.index:
                            row_ser["الماركة"] = rrow.get("الماركة", "")
                        rows_out.append(row_ser)
            if rows_out:
                final_cmp = pd.DataFrame(rows_out)
                st.session_state.cmp_export_df = _prepare_salla_product_df_for_export(final_cmp)
                st.success(f"✅ {len(final_cmp)} منتج في القائمة النهائية")
            else:
                st.warning("لا توجد صفوف معتمدة.")

        if getattr(st.session_state, "cmp_export_df", None) is not None:
            fe = datetime.now().strftime("%Y%m%d_%H%M")
            st.download_button(
                "📥 تنزيل Excel",
                export_product_xlsx(st.session_state.cmp_export_df),
                f"mahwous_after_compare_{fe}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
                key="cmp_dl_x",
            )

        if st.button("🔄 إعادة ضبط المقارنة", key="cmp_reset"):
            st.session_state.cmp_results = None
            st.session_state.cmp_approved = {}
            st.session_state.cmp_edit_name = {}
            if hasattr(st.session_state, "cmp_export_df"):
                del st.session_state["cmp_export_df"]
            for _k in list(st.session_state.keys()):
                if isinstance(_k, str) and _k.startswith("cmp_edit_"):
                    del st.session_state[_k]
            st.rerun()

    elif st.session_state.cmp_new_df is None or st.session_state.cmp_store_df is None:
        st.markdown("""<div class="upload-zone"><div class="uz-icon">🔀</div>
        <div class="uz-title">ارفع ملف المنتجات الجديدة وملف المتجر</div>
        </div>""", unsafe_allow_html=True)


def render_store_audit_tab():
    _legacy_init_state()
    _legacy_inject_css()
    st.markdown("""<div class="al-info">
    ارفع ملف المتجر الأساسي (بتنسيق سلة). فحص النواقص وإصلاح تلقائي اختياري.
    </div>""", unsafe_allow_html=True)

    st.markdown("""<div class="sec-title"><div class="bar"></div><h3>رفع ملف المتجر</h3></div>""",
                unsafe_allow_html=True)

    up_audit = st.file_uploader("ارفع ملف المتجر الأساسي (CSV / Excel)",
                                type=["csv", "xlsx", "xls"], key="sa_audit_up")
    if up_audit:
        df_audit_raw = read_file(up_audit, salla_2row=True)
        if df_audit_raw.empty:
            df_audit_raw = read_file(up_audit, salla_2row=False)
        if not df_audit_raw.empty:
            st.session_state.audit_df = df_audit_raw
            st.success(f"✅ {len(df_audit_raw):,} منتج في الملف")
            st.rerun()

    if st.session_state.audit_df is not None:
        audit_df = st.session_state.audit_df

        st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
        <h3>تعيين الأعمدة</h3></div>""", unsafe_allow_html=True)

        NONE_A = "— لا يوجد —"
        a_opts = [NONE_A] + list(audit_df.columns)

        def agi(kws):
            g = auto_guess_col(audit_df.columns, kws, audit_df)
            return a_opts.index(g) if g in a_opts else 0

        a1, a2, a3, a4, a5, a6 = st.columns(6)
        with a1:
            a_no = st.selectbox("No.", a_opts, index=agi(["no.", "no", "رقم", "id"]), key="sa_a_no")
        with a2:
            a_nm = st.selectbox("اسم المنتج", a_opts, index=agi(["اسم", "name", "منتج"]), key="sa_a_nm")
        with a3:
            a_img = st.selectbox("الصورة", a_opts, index=agi(["صورة", "image", "img"]), key="sa_a_img")
        with a4:
            a_cat = st.selectbox("التصنيف", a_opts, index=agi(["تصنيف", "category", "قسم"]), key="sa_a_cat")
        with a5:
            a_br = st.selectbox("الماركة", a_opts, index=agi(["ماركة", "brand", "علامة"]), key="sa_a_br")
        with a6:
            a_desc = st.selectbox("الوصف", a_opts, index=agi(["وصف", "description", "desc"]), key="sa_a_desc")

        a7, a8, a9 = st.columns(3)
        with a7:
            a_pr = st.selectbox("السعر", a_opts, index=agi(["سعر", "price"]), key="sa_a_pr")
        with a8:
            a_sku = st.selectbox("SKU", a_opts, index=agi(["sku", "رمز", "barcode"]), key="sa_a_sku")
        with a9:
            a_stat = st.selectbox(
                "حالة المنتج", a_opts, index=agi(["حالة المنتج", "حالة", "status", "مرئي", "متاح"]), key="sa_a_stat")

        a10, a11, a12 = st.columns(3)
        with a10:
            a_cost = st.selectbox(
                "سعر التكلفة", a_opts,
                index=agi(["سعر التكلفة", "تكلفة", "cost"]),
                key="sa_a_cost",
            )
        with a11:
            a_sale = st.selectbox(
                "السعر المخفض", a_opts,
                index=agi(["السعر المخفض", "مخفض", "sale", "discount"]),
                key="sa_a_sale",
            )
        with a12:
            a_tax = st.selectbox(
                "خاضع للضريبة ؟", a_opts,
                index=agi(["خاضع للضريبة ؟", "خاضع للضريبة", "ضريبة", "tax"]),
                key="sa_a_tax",
            )

        min_margin_pct = st.number_input(
            "الحد الأدنى للهامش الربحي (%)",
            min_value=0.0, max_value=90.0, value=15.0, step=0.5,
            key="sa_min_margin_pct",
        )

        if st.button("🔍 فحص الملف الآن", type="primary", key="sa_run_audit"):
            issues = []
            prog_bar = st.progress(0, text="جاري فحص المنتجات...")
            if a_pr == NONE_A or a_cost == NONE_A:
                st.error("حدد عمود 'السعر' وعمود 'سعر التكلفة' لتدقيق الهامش.")
                st.stop()
            total = len(audit_df)
            for i, row in audit_df.iterrows():
                if i % 10 == 0:
                    prog_bar.progress(int((i / max(total, 1)) * 100), text=f"فحص: {i}/{total}")
                row_issues = []
                name = str(row.get(a_nm, "") or "").strip() if a_nm != NONE_A else ""
                if not name or name == "nan":
                    continue

                if a_img != NONE_A and not str(row.get(a_img, "") or "").strip():
                    row_issues.append("بدون صورة")
                if a_cat != NONE_A and not str(row.get(a_cat, "") or "").strip():
                    row_issues.append("بدون تصنيف")
                if a_br != NONE_A and not str(row.get(a_br, "") or "").strip():
                    row_issues.append("بدون ماركة")

                desc_val = str(row.get(a_desc, "") or "").strip() if a_desc != NONE_A else ""
                if not desc_val or desc_val == "nan" or len(desc_val) < 20:
                    row_issues.append("بدون وصف")
                elif ("تستر" in name.lower() or "tester" in name.lower()) and "تستر" not in desc_val and "tester" not in desc_val.lower():
                    row_issues.append("وصف التستر غير صحيح")

                if a_pr != NONE_A:
                    pr_raw = str(row.get(a_pr, "") or "").strip()
                    if pr_raw in ["0", "nan", ""]:
                        row_issues.append("بدون سعر")
                    else:
                        _ok_pr, _pv = parse_price_numeric(pr_raw)
                        if not _ok_pr or _pv <= 0:
                            row_issues.append("بدون سعر")

                ok_pr = False
                pv = 0.0
                ok_cost = False
                cost_v = 0.0
                if a_pr != NONE_A:
                    pr_raw = str(row.get(a_pr, "") or "").strip()
                    ok_pr, pv = parse_price_numeric(pr_raw)
                if a_cost != NONE_A:
                    cost_raw = str(row.get(a_cost, "") or "").strip()
                    ok_cost, cost_v = parse_price_numeric(cost_raw)

                if a_cost != NONE_A and (not ok_cost or cost_v <= 0):
                    row_issues.append("بدون سعر تكلفة")

                if a_pr != NONE_A and a_cost != NONE_A and ok_pr and ok_cost and pv > 0 and cost_v > 0:
                    margin_pct = (pv - cost_v) / cost_v * 100.0
                    if pv < cost_v:
                        row_issues.append("السعر أقل من التكلفة")
                    elif margin_pct < float(min_margin_pct):
                        row_issues.append(f"هامش منخفض (<{min_margin_pct}%)")

                if row_issues:
                    no_cell = ""
                    sku_cell = ""
                    if a_no != NONE_A:
                        no_cell = str(row.get(a_no, "") or "").strip()
                    if a_sku != NONE_A:
                        sku_cell = str(row.get(a_sku, "") or "").strip()
                    stat_cell = "مرئي"
                    if a_stat != NONE_A:
                        stat_cell = str(row.get(a_stat, "") or "").strip() or "مرئي"

                    tax_cell = "نعم"
                    if a_tax != NONE_A:
                        tax_cell = str(row.get(a_tax, "") or "").strip()
                        if not tax_cell or tax_cell.lower() in ("nan", "none"):
                            tax_cell = "نعم"
                        else:
                            tl = tax_cell.lower()
                            if tl in ("نعم", "yes", "true", "1"):
                                tax_cell = "نعم"
                            elif tl in ("لا", "no", "false", "0"):
                                tax_cell = "لا"
                            else:
                                tax_cell = str(row.get(a_tax, "") or "").strip()

                    issues.append({
                        "No.": no_cell,
                        "النوع ": "منتج",
                        "أسم المنتج": name,
                        "الماركة": str(row.get(a_br, "") or "") if a_br != NONE_A else "",
                        "تصنيف المنتج": str(row.get(a_cat, "") or "") if a_cat != NONE_A else "",
                        "صورة المنتج": str(row.get(a_img, "") or "") if a_img != NONE_A else "",
                        "وصف صورة المنتج": name,
                        "نوع المنتج": "منتج جاهز",
                        "سعر المنتج": normalize_price_digits(row.get(a_pr, "")) if a_pr != NONE_A else "",
                        "سعر التكلفة": normalize_price_digits(row.get(a_cost, "")) if a_cost != NONE_A else "",
                        "السعر المخفض": normalize_price_digits(row.get(a_sale, "")) if a_sale != NONE_A else "",
                        "الوصف": desc_val,
                        "هل يتطلب شحن؟": "نعم",
                        "رمز المنتج sku": sku_cell,
                        "الوزن": "0.2",
                        "وحدة الوزن": "kg",
                        "حالة المنتج": stat_cell,
                        "خاضع للضريبة ؟": tax_cell,
                        "اقصي كمية لكل عميل": "0",
                        "تثبيت المنتج": "لا",
                        "اضافة صورة عند الطلب": "لا",
                        "_issues": " | ".join(row_issues),
                        "_idx": i,
                    })
            prog_bar.progress(100, text="اكتمل الفحص!")
            st.session_state.audit_results = pd.DataFrame(issues) if issues else pd.DataFrame()
            st.rerun()

        if st.session_state.audit_results is not None:
            audit_res = st.session_state.audit_results

            if audit_res.empty:
                st.success("✅ الملف مكتمل — لا توجد منتجات تحتاج معالجة!")
            else:
                st.dataframe(
                    audit_res[[c for c in ["No.", "أسم المنتج", "الماركة", "تصنيف المنتج", "_issues"] if c in audit_res.columns]],
                    width="stretch",
                )
                date_str = datetime.now().strftime("%Y%m%d_%H%M")
                st.download_button(
                    "📥 تصدير تدقيق — CSV",
                    export_product_csv(audit_res),
                    f"price_audit_{date_str}.csv",
                    "text/csv",
                    width="stretch",
                    key="sa_price_audit_dl_csv",
                )

                st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
                <h3>🛠️ الإصلاح التلقائي</h3></div>""", unsafe_allow_html=True)

                if st.button("🚀 بدء إصلاح النواقص أوتوماتيكياً", type="primary",
                             key="sa_start_auto_fix", width="stretch"):
                    if not st.session_state.api_key:
                        st.error("⚠️ يرجى إضافة مفتاح Claude API (Anthropic) في secrets أو الإعدادات.")
                    else:
                        fix_prog = st.progress(0)
                        fixed_rows = []
                        total_fix = len(audit_res)
                        for fix_i, (_, f_row) in enumerate(audit_res.iterrows()):
                            fix_prog.progress(int((fix_i / max(total_fix, 1)) * 100))
                            pname = f_row["أسم المنتج"]
                            iss = f_row["_issues"]
                            attrs = extract_product_attrs(pname)
                            size_f = attrs.get("size") or 0
                            size = f"{int(size_f) if size_f == int(size_f) else size_f} مل" if size_f else "100 مل"
                            size = _normalize_product_size_ml(size) or size
                            conc_ar = "أو دو بارفيوم"
                            is_t = "تستر" in str(attrs.get("type", ""))
                            gender = ("للنساء" if any(w in pname.lower() for w in ["نسائ", "women"]) else
                                      "للرجال" if any(w in pname.lower() for w in ["رجال", "men"]) else "للجنسين")
                            brand_dict = match_brand(pname)
                            if not brand_dict.get("name") and f_row["الماركة"]:
                                brand_dict = match_brand(f_row["الماركة"])
                            if "بدون ماركة" in iss or not brand_dict.get("name"):
                                if not brand_dict.get("name"):
                                    extracted_b = clean_brand_name(pname.split()[0] if pname.split() else "")
                                    if extracted_b:
                                        brand_dict = generate_new_brand(extracted_b)
                            f_row["الماركة"] = brand_dict.get("name", f_row["الماركة"])
                            f_row["أسم المنتج"] = standardize_product_name(pname, f_row["الماركة"])
                            pname = f_row["أسم المنتج"]
                            if "بدون تصنيف" in iss or not f_row["تصنيف المنتج"]:
                                f_row["تصنيف المنتج"] = ("العطور > تستر" if is_t else match_category(pname, gender))
                            if "بدون صورة" in iss or not f_row["صورة المنتج"]:
                                f_row["صورة المنتج"] = fetch_image(pname, is_t)
                            if "بدون وصف" in iss or "وصف التستر غير صحيح" in iss:
                                f_row["الوصف"] = ai_generate(
                                    pname, is_t, brand_dict, size, gender, conc_ar)
                            no_seo = str(f_row.get("No.", "") or fix_i + 1).strip()
                            seo_data = gen_seo(
                                pname, brand_dict, size, is_t, gender,
                                sku_suffix=f"V-{no_seo}",
                                type_hint=str(attrs.get("type", "") or ""),
                            )
                            f_row["وصف صورة المنتج"] = seo_data["alt"]
                            final_dict = {col: "" for col in SALLA_COLS}
                            for col in SALLA_COLS:
                                if col in f_row:
                                    final_dict[col] = str(f_row[col])
                            final_dict["النوع "] = "منتج"
                            fixed_rows.append(final_dict)
                        fix_prog.progress(100)
                        st.session_state.audit_fixed_df = pd.DataFrame(fixed_rows, columns=SALLA_COLS)

                if "audit_fixed_df" in st.session_state and st.session_state.audit_fixed_df is not None:
                    date_str = datetime.now().strftime("%Y%m%d_%H%M")
                    aud_e1, aud_e2 = st.columns(2)
                    with aud_e1:
                        st.download_button(
                            f"📥 الملف المُصلح — Excel ({len(st.session_state.audit_fixed_df)})",
                            export_product_xlsx(st.session_state.audit_fixed_df),
                            f"تحديث_منتجات_{date_str}.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            width="stretch", key="sa_dl_audit_fix_x")
                    with aud_e2:
                        st.download_button(
                            f"📥 الملف المُصلح — CSV ({len(st.session_state.audit_fixed_df)})",
                            export_product_csv(st.session_state.audit_fixed_df),
                            f"تحديث_منتجات_{date_str}.csv",
                            "text/csv", width="stretch", key="sa_dl_audit_fix_c")

                    st.info("للمتابعة في المسار الآلي الكامل استخدم الأقسام الرئيسية في تطبيق v26.")

                if st.button("🔄 إعادة الفحص", key="sa_reset_audit"):
                    st.session_state.audit_results = None
                    if "audit_fixed_df" in st.session_state:
                        del st.session_state["audit_fixed_df"]
                    st.rerun()

    else:
        st.markdown("""
        <div class="upload-zone">
          <div class="uz-icon">🏪</div>
          <div class="uz-title">ارفع ملف المتجر الأساسي للبدء</div>
        </div>
        """, unsafe_allow_html=True)


def render_seo_processor_tab():
    _legacy_init_state()
    _legacy_inject_css()
    st.markdown("## معالج الـ SEO (Salla)")
    st.markdown("""<div class="al-info">
    ارفع ملف منتجات سلة كامل (Excel أو CSV). يُولّد الحقول الناقصة عبر Anthropic.
    </div>""", unsafe_allow_html=True)

    up_seo = st.file_uploader(
        "ارفع ملف منتجات سلة SEO",
        type=["csv", "xlsx", "xls", "xlsm"],
        key="seo_proc_tab_uploader",
    )

    if up_seo:
        df_seo = read_file(up_seo, salla_2row=True)
        if df_seo.empty:
            df_seo = read_file(up_seo, salla_2row=False)
        if not df_seo.empty:
            st.session_state.seo_proc_tab_input_df = df_seo
            st.success(f"✅ تم تحميل {len(df_seo):,} صف")
            st.rerun()

    input_df = getattr(st.session_state, "seo_proc_tab_input_df", None)
    if input_df is None:
        return

    sdf = input_df
    name_col = "اسم المنتج (غير قابل للتعديل)"
    url_col = "رابط مخصص للمنتج (SEO Page URL)"
    title_col = "عنوان صفحة المنتج (SEO Page Title)"
    desc_col = "وصف صفحة المنتج (SEO Page Description)"

    for col in (name_col, url_col, title_col, desc_col):
        if col not in sdf.columns:
            st.error(f"الملف لا يحتوي العمود المطلوب: {col}")
            return

    with st.expander("👀 معاينة الملف", expanded=False):
        st.dataframe(sdf.head(12), width="stretch")

    if st.button("🚀 بدء المعالجة", type="primary", key="seo_proc_tab_run", width="stretch"):
        if not st.session_state.api_key:
            st.error("أضف مفتاح Anthropic API من secrets أو الإعدادات.")
            st.stop()

        work = sdf.copy()

        def cell_is_empty(v) -> bool:
            s = str(v or "").strip()
            if not s:
                return True
            return s.lower() in ("nan", "none")

        prog = st.progress(0, text="جاري معالجة الصفوف...")
        total = max(len(work), 1)

        for i in range(len(work)):
            row_dict = work.iloc[i].to_dict()
            prod_name = str(row_dict.get(name_col, "") or "").strip()
            if not prod_name or prod_name.lower() in ("nan", "none"):
                continue

            missing_fields: list[str] = []
            if cell_is_empty(row_dict.get(url_col, "")):
                missing_fields.append("url")
            if cell_is_empty(row_dict.get(title_col, "")):
                missing_fields.append("title")
            if cell_is_empty(row_dict.get(desc_col, "")):
                missing_fields.append("desc")

            if missing_fields:
                out = generate_seo_data_ai(prod_name, missing_fields)
                if "url_slug" in out and "url" in missing_fields:
                    work.at[i, url_col] = out["url_slug"]
                if "page_title" in out and "title" in missing_fields:
                    work.at[i, title_col] = out["page_title"]
                if "meta_description" in out and "desc" in missing_fields:
                    work.at[i, desc_col] = out["meta_description"]

            prog.progress(int((i + 1) / total * 100), text=f"معالجة: {i+1}/{total}")

        st.session_state.seo_proc_tab_output_df = work[SALLA_SEO_COLS].copy()
        st.success("✅ تم تحديث ملف SEO بنجاح")
        st.rerun()

    output_df = getattr(st.session_state, "seo_proc_tab_output_df", None)
    if output_df is not None and not output_df.empty:
        st.markdown("""<div class="sec-title"><div class="bar"></div>
        <h3>نتيجة التوليد (تنسيق سلة SEO)</h3></div>""", unsafe_allow_html=True)
        edited_df = st.data_editor(
            output_df.fillna(""),
            width="stretch",
            num_rows="dynamic",
            key="seo_proc_tab_editor",
        )
        st.session_state.seo_proc_tab_output_df = edited_df

        csv_str = edited_df.to_csv(index=False, encoding="utf-8-sig")
        st.download_button(
            "📥 تصدير SEO — CSV (محدث)",
            csv_str,
            "مهووس_SEO_محدث.csv",
            "text/csv",
            width="stretch",
            key="seo_proc_tab_dl",
        )



def render_legacy_dashboard():
    """لوحة الأدوات الوراثية — تبويبات المقارنة، التدقيق، SEO."""
    _legacy_inject_css()
    _legacy_init_state()

    c1, c2 = st.columns([4, 1])
    with c1:
        st.title("🛠️ التدقيق والتحسين")
    with c2:
        if st.button("← العودة للتطبيق", key="legacy_back_main", width="stretch"):
            st.session_state.legacy_tools_mode = False
            st.rerun()

    st.caption("أدوات v11 المعزولة — تتطلب مفتاح Anthropic في الإعدادات أو secrets.")

    tab1, tab2, tab3 = st.tabs(["🔀 المقارنة", "🏪 مدقق المتجر", "🔍 معالج السيو"])
    with tab1:
        render_compare_tab()
    with tab2:
        render_store_audit_tab()
    with tab3:
        render_seo_processor_tab()
